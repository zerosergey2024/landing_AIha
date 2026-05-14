import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "leads.db"
EXPORT_PATH = BASE_DIR / "exports" / "leads_export.xlsx"


def export_leads_to_xlsx():
    EXPORT_PATH.parent.mkdir(exist_ok=True)

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT
                id,
                created_at,
                updated_at,
                source,
                name,
                phone,
                company,
                message,
                industry,
                process,
                ai_type,
                effect,
                priority,
                status,
                manager_comment
            FROM leads
            ORDER BY id DESC
            """
        ).fetchall()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Заявки"

    headers = [
        "ID",
        "Дата создания",
        "Дата обновления",
        "Источник",
        "Имя",
        "Телефон",
        "Компания",
        "Задача",
        "Отрасль",
        "Процесс",
        "AI-сценарий",
        "Эффект",
        "Приоритет",
        "Статус",
        "Комментарий менеджера",
    ]

    sheet.append(headers)

    for row in rows:
        sheet.append([
            row["id"],
            row["created_at"],
            row["updated_at"],
            row["source"],
            row["name"],
            row["phone"],
            row["company"],
            row["message"],
            row["industry"],
            row["process"],
            row["ai_type"],
            row["effect"],
            row["priority"],
            row["status"],
            row["manager_comment"],
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 8,
        "B": 22,
        "C": 22,
        "D": 18,
        "E": 22,
        "F": 22,
        "G": 26,
        "H": 55,
        "I": 24,
        "J": 30,
        "K": 28,
        "L": 36,
        "M": 16,
        "N": 16,
        "O": 38,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:O{sheet.max_row}"

    workbook.save(EXPORT_PATH)

    print(f"XLSX экспорт готов: {EXPORT_PATH}")


if __name__ == "__main__":
    export_leads_to_xlsx()