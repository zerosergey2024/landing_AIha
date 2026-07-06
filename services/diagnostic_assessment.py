from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from db import get_db_connection
from services.ai_agent import run_agent_with_prompt
from services.diagnostic_normalizer import normalize_input_pack
from services.diagnostics import (
    DIAGNOSTIC_STATUS_D001_COMPLETED,
    DIAGNOSTIC_STATUS_D001_RUNNING,
    save_d001_result,
    update_diagnostic_status,
)


D001_AGENT_PROMPT_NAME = "diagnostic_assessment"
MIN_INDUSTRIAL_FILLED_FIELDS = 20

MAX_ATTACHMENT_SIZE_MB = 50
MAX_TOTAL_ATTACHMENTS_SIZE_MB = 200
MAX_ATTACHMENTS_PER_BRIEF = 10

MAX_TABLE_PREVIEW_ROWS = 20
MAX_TABLE_PREVIEW_COLUMNS = 60
MAX_WORKSHEET_PREVIEW_COUNT = 5

MAX_DOCUMENT_TEXT_CHARS = 30_000
MAX_PDF_PREVIEW_PAGES = 8

MAX_SINGLE_ATTACHMENT_PROMPT_CHARS = 25_000
MAX_TOTAL_ATTACHMENT_PROMPT_CHARS = 80_000

SUPPORTED_ATTACHMENT_PREVIEW_SUFFIXES = {
    ".xlsx",
    ".xlsm",
    ".xls",
    ".csv",
    ".docx",
    ".pdf",
}

BYTES_IN_MB = 1024 * 1024
MAX_ATTACHMENT_SIZE_BYTES = MAX_ATTACHMENT_SIZE_MB * BYTES_IN_MB
MAX_TOTAL_ATTACHMENTS_SIZE_BYTES = MAX_TOTAL_ATTACHMENTS_SIZE_MB * BYTES_IN_MB


def _safe_string(value: Any, max_len: int = 240) -> str:
    if value is None:
        return ""

    text = str(value).strip()

    if len(text) > max_len:
        return text[:max_len] + "..."

    return text


def _truncate_text(value: Any, max_chars: int) -> tuple[str, bool, int]:
    text = str(value or "")
    original_length = len(text)

    if original_length <= max_chars:
        return text, False, original_length

    return (
        text[:max_chars]
        + "\n\n[TRUNCATED: attachment preview limited to "
        + str(max_chars)
        + " characters for consulting-grade diagnostic review]",
        True,
        original_length,
    )


def _limit_text(value: Any, max_chars: int) -> str:
    limited_text, _, _ = _truncate_text(value, max_chars)
    return limited_text


def _bytes_to_mb(size_bytes: int | None) -> float | None:
    if size_bytes is None:
        return None

    return round(size_bytes / BYTES_IN_MB, 3)


def _build_attachment_limits_summary() -> dict[str, Any]:
    return {
        "mode": "limited_consulting_grade_review",
        "stored_file": "full_file_is_stored",
        "prompt_usage": "compact_preview_and_evidence_summary_only",
        "upload_limits": {
            "max_attachment_size_mb": MAX_ATTACHMENT_SIZE_MB,
            "max_total_attachments_size_mb": MAX_TOTAL_ATTACHMENTS_SIZE_MB,
            "max_attachments_per_brief": MAX_ATTACHMENTS_PER_BRIEF,
        },
        "read_limits": {
            "max_table_preview_rows": MAX_TABLE_PREVIEW_ROWS,
            "max_table_preview_columns": MAX_TABLE_PREVIEW_COLUMNS,
            "max_worksheet_preview_count": MAX_WORKSHEET_PREVIEW_COUNT,
            "max_document_text_chars": MAX_DOCUMENT_TEXT_CHARS,
            "max_pdf_preview_pages": MAX_PDF_PREVIEW_PAGES,
            "max_single_attachment_prompt_chars": MAX_SINGLE_ATTACHMENT_PROMPT_CHARS,
            "max_total_attachment_prompt_chars": MAX_TOTAL_ATTACHMENT_PROMPT_CHARS,
        },
        "not_in_scope": [
            "full BI analytics",
            "deep sensor analytics",
            "production model training",
            "final ROI calculation without baseline",
            "legal or audit-grade conclusions",
        ],
    }


def _load_json_object(value: str | None, label: str) -> dict[str, Any]:
    if not value:
        return {}

    try:
        parsed = json.loads(value)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON in {label}: {exc}") from exc

    if not isinstance(parsed, dict):
        raise ValueError(f"{label} must contain a JSON object")

    return parsed


def _count_non_empty(value: Any) -> int:
    if isinstance(value, dict):
        return sum(_count_non_empty(item) for item in value.values())

    if isinstance(value, list):
        return sum(1 for item in value if str(item).strip())

    return 1 if value is not None and str(value).strip() else 0


def _sanitize_diagnostic_run(row: Any) -> dict[str, Any]:
    diagnostic_run = dict(row)

    keys_to_remove = {
        "input_pack_token",
        "d001_result",
        "d002_result",
        "d003_result",
        "d004_result",
        "final_result",
        "final_output",
        "diagnostic_report",
        "commercial_proposal",
    }

    for key in keys_to_remove:
        diagnostic_run.pop(key, None)

    return diagnostic_run


def _validate_raw_payload_for_d001(
    input_pack_id: int,
    raw_payload: dict[str, Any],
) -> None:
    if not raw_payload:
        raise ValueError(
            f"raw_payload is empty for diagnostic input pack id={input_pack_id}"
        )

    brief_type = raw_payload.get("brief_type")

    if brief_type != "industrial_ai":
        return

    industrial_ai = raw_payload.get("industrial_ai")

    if not isinstance(industrial_ai, dict):
        raise ValueError(
            f"industrial_ai object is missing in raw_payload for input_pack id={input_pack_id}"
        )

    filled_fields = _count_non_empty(industrial_ai)

    if filled_fields < MIN_INDUSTRIAL_FILLED_FIELDS:
        raise ValueError(
            "Industrial AI Brief is too sparse for D-001: "
            f"input_pack_id={input_pack_id}, "
            f"industrial_non_empty={filled_fields}. "
            "Submit a filled Industrial AI Brief before running D-001."
        )


def _get_diagnostic_run_row(diagnostic_run_id: int) -> Any:
    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM diagnostic_runs
            WHERE id = ?
            """,
            (diagnostic_run_id,),
        ).fetchone()

    if row is None:
        raise ValueError(f"Diagnostic run not found: {diagnostic_run_id}")

    return row


def _get_latest_input_pack_row(diagnostic_run_id: int) -> Any:
    _get_diagnostic_run_row(diagnostic_run_id)

    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM diagnostic_input_packs
            WHERE diagnostic_run_id = ?
              AND is_active = 1
              AND brief_type = 'industrial_ai'
              AND raw_payload IS NOT NULL
              AND TRIM(raw_payload) != ''
            ORDER BY
                COALESCE(updated_at, created_at) DESC,
                id DESC
            LIMIT 1
            """,
            (diagnostic_run_id,),
        ).fetchone()

        if row is None:
            row = conn.execute(
                """
                SELECT *
                FROM diagnostic_input_packs
                WHERE diagnostic_run_id = ?
                  AND is_active = 1
                  AND brief_type = 'diagnostic_input_pack'
                  AND raw_payload IS NOT NULL
                  AND TRIM(raw_payload) != ''
                ORDER BY
                    COALESCE(updated_at, created_at) DESC,
                    id DESC
                LIMIT 1
                """,
                (diagnostic_run_id,),
            ).fetchone()

    if row is None:
        raise ValueError(
            "Active diagnostic input pack with raw_payload not found "
            f"for diagnostic_run_id={diagnostic_run_id}"
        )

    return row


def _get_attachments_for_input_pack(input_pack_id: int) -> list[dict[str, Any]]:
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                id,
                input_pack_id,
                file_type,
                original_filename,
                stored_filename,
                file_path,
                uploaded_at
            FROM diagnostic_attachments
            WHERE input_pack_id = ?
            ORDER BY id ASC
            """,
            (input_pack_id,),
        ).fetchall()

    return [dict(row) for row in rows]


def validate_d001_input(diagnostic_run_id: int) -> None:
    input_pack = _get_latest_input_pack_row(diagnostic_run_id)
    raw_payload = _load_json_object(input_pack["raw_payload"], "raw_payload")

    _validate_raw_payload_for_d001(
        input_pack_id=int(input_pack["id"]),
        raw_payload=raw_payload,
    )


def _resolve_existing_path(file_path: str | None) -> Path:
    raw_path = str(file_path or "").strip()

    if not raw_path:
        return Path("")

    path = Path(raw_path)

    if path.exists():
        return path

    candidates = [
        Path.cwd() / raw_path,
        Path.cwd() / raw_path.lstrip("/\\"),
    ]

    for candidate in candidates:
        if candidate.exists():
            return candidate

    return path


def _get_file_size_bytes(path: Path) -> int | None:
    try:
        if path.exists() and path.is_file():
            return path.stat().st_size
    except OSError:
        return None

    return None


def _build_attachment_inventory(
    attachments: list[dict[str, Any]],
) -> dict[str, Any]:
    files: list[dict[str, Any]] = []
    total_size_bytes = 0
    known_size_count = 0

    for index, attachment in enumerate(attachments):
        file_path = attachment.get("file_path") or ""
        resolved_path = _resolve_existing_path(file_path)
        exists = bool(file_path) and resolved_path.exists()
        size_bytes = _get_file_size_bytes(resolved_path)

        if size_bytes is not None:
            known_size_count += 1
            total_size_bytes += size_bytes

        files.append(
            {
                "attachment_id": attachment.get("id"),
                "original_filename": attachment.get("original_filename"),
                "stored_filename": attachment.get("stored_filename"),
                "file_type": attachment.get("file_type"),
                "uploaded_at": attachment.get("uploaded_at"),
                "file_path": file_path,
                "exists": exists,
                "size_bytes": size_bytes,
                "size_mb": _bytes_to_mb(size_bytes),
                "within_single_file_limit": (
                    size_bytes is None or size_bytes <= MAX_ATTACHMENT_SIZE_BYTES
                ),
                "within_count_limit": index < MAX_ATTACHMENTS_PER_BRIEF,
            }
        )

    limit_violations: list[str] = []

    if len(attachments) > MAX_ATTACHMENTS_PER_BRIEF:
        limit_violations.append("attachments_count_exceeds_limit")

    if total_size_bytes > MAX_TOTAL_ATTACHMENTS_SIZE_BYTES:
        limit_violations.append("known_total_attachment_size_exceeds_limit")

    return {
        "files_count": len(attachments),
        "known_size_count": known_size_count,
        "total_known_size_bytes": total_size_bytes,
        "total_known_size_mb": _bytes_to_mb(total_size_bytes),
        "limits": _build_attachment_limits_summary()["upload_limits"],
        "limit_violations": limit_violations,
        "files": files,
    }


def _make_attachment_item(
    attachment: dict[str, Any],
    resolved_path: Path,
    suffix: str,
) -> dict[str, Any]:
    size_bytes = _get_file_size_bytes(resolved_path)

    return {
        "attachment_id": attachment.get("id"),
        "input_pack_id": attachment.get("input_pack_id"),
        "original_filename": attachment.get("original_filename") or "",
        "stored_filename": attachment.get("stored_filename"),
        "file_type": attachment.get("file_type"),
        "uploaded_at": attachment.get("uploaded_at"),
        "file_path": attachment.get("file_path") or "",
        "resolved_file_path": str(resolved_path),
        "suffix": suffix,
        "exists": resolved_path.exists(),
        "file_size_bytes": size_bytes,
        "file_size_mb": _bytes_to_mb(size_bytes),
        "preview": None,
        "detected_signals": None,
    }


def _read_csv_preview(
    file_path: str,
    *,
    max_rows: int = MAX_TABLE_PREVIEW_ROWS,
    max_columns: int = MAX_TABLE_PREVIEW_COLUMNS,
) -> dict[str, Any]:
    path = Path(file_path)

    result: dict[str, Any] = {
        "kind": "csv_table_preview",
        "file_path": str(path),
        "exists": path.exists(),
        "columns": [],
        "columns_count": 0,
        "sample_rows": [],
        "rows_count": 0,
        "row_count_previewed": 0,
        "encoding": None,
        "limits_applied": {
            "max_preview_rows": max_rows,
            "max_preview_columns": max_columns,
            "rows_truncated": False,
            "columns_truncated": False,
        },
        "error": None,
    }

    if not path.exists():
        result["error"] = "file_not_found"
        return result

    last_error: Exception | None = None

    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                reader = csv.DictReader(file)
                all_columns = [
                    _safe_string(column)
                    for column in (reader.fieldnames or [])
                    if column is not None
                ]

                selected_columns = all_columns[:max_columns]
                sample_rows: list[dict[str, str]] = []
                rows_count = 0

                for row in reader:
                    rows_count += 1

                    if len(sample_rows) >= max_rows:
                        continue

                    row_dict: dict[str, str] = {}

                    for column in selected_columns:
                        row_dict[column] = _safe_string(row.get(column))

                    if any(value.strip() for value in row_dict.values()):
                        sample_rows.append(row_dict)

                result["columns"] = selected_columns
                result["columns_count"] = len(all_columns)
                result["sample_rows"] = sample_rows
                result["rows_count"] = rows_count
                result["row_count_previewed"] = len(sample_rows)
                result["encoding"] = encoding
                result["limits_applied"] = {
                    "max_preview_rows": max_rows,
                    "max_preview_columns": max_columns,
                    "rows_truncated": rows_count > max_rows,
                    "columns_truncated": len(all_columns) > max_columns,
                }

                return result

        except Exception as exc:
            last_error = exc

    result["error"] = str(last_error) if last_error else "csv_read_failed"
    return result


def _read_xlsx_preview(
    file_path: str,
    *,
    max_rows: int = MAX_TABLE_PREVIEW_ROWS,
    max_columns: int = MAX_TABLE_PREVIEW_COLUMNS,
) -> dict[str, Any]:
    path = Path(file_path)

    result: dict[str, Any] = {
        "kind": "xlsx_table_preview",
        "file_path": str(path),
        "exists": path.exists(),
        "sheets": [],
        "sheets_count": 0,
        "limits_applied": {
            "max_preview_rows": max_rows,
            "max_preview_columns": max_columns,
            "max_worksheet_preview_count": MAX_WORKSHEET_PREVIEW_COUNT,
            "sheets_truncated": False,
        },
        "error": None,
    }

    if not path.exists():
        result["error"] = "file_not_found"
        return result

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        result["error"] = f"openpyxl_not_available: {exc}"
        return result

    def normalize_cell(value: Any) -> str:
        return _safe_string(value, max_len=120).strip()

    def header_score(row_values: list[Any]) -> int:
        cells = [
            normalize_cell(value).lower()
            for value in row_values
            if normalize_cell(value)
        ]

        if not cells:
            return 0

        known_header_terms = {
            "event_id",
            "downtime_event_id",
            "event_date",
            "line_id",
            "line_name",
            "equipment_id",
            "equipment_name",
            "downtime_start",
            "downtime_end",
            "start_time",
            "end_time",
            "duration_min",
            "duration_minutes",
            "downtime_minutes",
            "reason_code",
            "reason_description",
            "downtime_reason",
            "source_system",
            "cost_per_hour_rub",
            "estimated_loss_rub",
            "comment",
            "corrective_action",
            "id события",
            "id линии",
            "id оборудования",
            "начало простоя",
            "конец простоя",
            "длительность",
            "причина простоя",
            "стоимость часа",
            "потери",
        }

        score = 0

        for cell in cells:
            if cell in known_header_terms:
                score += 5

            if any(term in cell for term in known_header_terms):
                score += 3

            if "_" in cell:
                score += 1

            if len(cell) <= 40:
                score += 1

        if len(cells) >= 5:
            score += 5

        if len(cells) == 1 and len(cells[0]) > 40:
            score -= 10

        return score

    workbook = None

    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
        )

        worksheets = workbook.worksheets
        result["sheets_count"] = len(worksheets)
        result["limits_applied"]["sheets_truncated"] = (
            len(worksheets) > MAX_WORKSHEET_PREVIEW_COUNT
        )

        for worksheet in worksheets[:MAX_WORKSHEET_PREVIEW_COUNT]:
            worksheet_max_row = worksheet.max_row or 0
            worksheet_max_column = worksheet.max_column or 0
            max_col_to_read = min(
                worksheet_max_column or max_columns,
                max_columns,
            )
            max_row_to_read = min(
                worksheet_max_row or 1,
                max_rows + 15,
            )

            preview_rows = list(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=max_row_to_read,
                    max_col=max_col_to_read,
                    values_only=True,
                )
            )

            if not preview_rows:
                result["sheets"].append(
                    {
                        "sheet_name": worksheet.title,
                        "max_row": worksheet_max_row,
                        "max_column": worksheet_max_column,
                        "header_row_index": None,
                        "columns": [],
                        "columns_count": worksheet_max_column,
                        "sample_rows": [],
                        "row_count_previewed": 0,
                        "limits_applied": {
                            "max_preview_rows": max_rows,
                            "max_preview_columns": max_columns,
                            "rows_truncated": False,
                            "columns_truncated": worksheet_max_column > max_columns,
                        },
                    }
                )
                continue

            best_header_index = 0
            best_score = -999

            for index, row_values in enumerate(preview_rows[:15]):
                score = header_score(list(row_values))

                if score > best_score:
                    best_score = score
                    best_header_index = index

            header_row = preview_rows[best_header_index]

            columns = [
                normalize_cell(value) if normalize_cell(value) else f"column_{index + 1}"
                for index, value in enumerate(header_row)
            ]

            columns = columns[:max_columns]
            data_rows = preview_rows[best_header_index + 1:]

            sample_rows: list[dict[str, str]] = []

            for row in data_rows:
                if len(sample_rows) >= max_rows:
                    break

                row_dict: dict[str, str] = {}

                for col_index, value in enumerate(row[: len(columns)]):
                    column_name = columns[col_index]
                    row_dict[column_name] = _safe_string(value)

                if any(value.strip() for value in row_dict.values()):
                    sample_rows.append(row_dict)

            data_rows_count_estimate = max(
                worksheet_max_row - best_header_index - 1,
                0,
            )

            result["sheets"].append(
                {
                    "sheet_name": worksheet.title,
                    "max_row": worksheet_max_row,
                    "max_column": worksheet_max_column,
                    "header_row_index": best_header_index + 1,
                    "header_detection_score": best_score,
                    "columns": columns,
                    "columns_count": worksheet_max_column,
                    "sample_rows": sample_rows,
                    "row_count_previewed": len(sample_rows),
                    "limits_applied": {
                        "max_preview_rows": max_rows,
                        "max_preview_columns": max_columns,
                        "rows_truncated": data_rows_count_estimate > max_rows,
                        "columns_truncated": worksheet_max_column > max_columns,
                    },
                }
            )

        return result

    except Exception as exc:
        result["error"] = str(exc)
        return result

    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def _read_xls_preview(
    file_path: str,
    *,
    max_rows: int = MAX_TABLE_PREVIEW_ROWS,
    max_columns: int = MAX_TABLE_PREVIEW_COLUMNS,
) -> dict[str, Any]:
    path = Path(file_path)

    result: dict[str, Any] = {
        "kind": "xls_table_preview",
        "file_path": str(path),
        "exists": path.exists(),
        "sheets": [],
        "limits_applied": {
            "max_preview_rows": max_rows,
            "max_preview_columns": max_columns,
            "rows_truncated": None,
            "columns_truncated": None,
        },
        "error": None,
    }

    if not path.exists():
        result["error"] = "file_not_found"
        return result

    try:
        import pandas as pd
    except Exception as exc:
        result["error"] = f"pandas_not_available: {exc}"
        return result

    try:
        sheets = pd.read_excel(
            path,
            sheet_name=None,
            nrows=max_rows,
        )

        for sheet_name, dataframe in list(sheets.items())[:MAX_WORKSHEET_PREVIEW_COUNT]:
            original_columns_count = len(dataframe.columns)
            limited_dataframe = dataframe.iloc[
                :max_rows,
                :max_columns,
            ]

            columns = [
                _safe_string(column)
                for column in limited_dataframe.columns.tolist()
            ]

            sample_rows: list[dict[str, str]] = []

            for _, row in limited_dataframe.iterrows():
                row_dict: dict[str, str] = {}

                for column in limited_dataframe.columns:
                    row_dict[_safe_string(column)] = _safe_string(row[column])

                if any(value.strip() for value in row_dict.values()):
                    sample_rows.append(row_dict)

            result["sheets"].append(
                {
                    "sheet_name": str(sheet_name),
                    "columns": columns,
                    "columns_count": original_columns_count,
                    "sample_rows": sample_rows,
                    "row_count_previewed": len(sample_rows),
                    "limits_applied": {
                        "max_preview_rows": max_rows,
                        "max_preview_columns": max_columns,
                        "rows_truncated": None,
                        "columns_truncated": original_columns_count > max_columns,
                    },
                }
            )

        result["limits_applied"]["columns_truncated"] = any(
            bool(sheet.get("limits_applied", {}).get("columns_truncated"))
            for sheet in result["sheets"]
        )

        return result

    except Exception as exc:
        result["error"] = str(exc)
        return result


def _read_docx_preview(
    file_path: str,
    *,
    max_chars: int = MAX_DOCUMENT_TEXT_CHARS,
) -> dict[str, Any]:
    path = Path(file_path)

    result: dict[str, Any] = {
        "kind": "docx_text_preview",
        "file_path": str(path),
        "exists": path.exists(),
        "text_excerpt": "",
        "paragraph_count_previewed": 0,
        "limits_applied": {
            "max_document_text_chars": max_chars,
            "text_truncated": False,
        },
        "error": None,
    }

    if not path.exists():
        result["error"] = "file_not_found"
        return result

    try:
        from docx import Document
    except Exception as exc:
        result["error"] = f"python_docx_not_available: {exc}"
        return result

    try:
        document = Document(path)
        parts: list[str] = []

        for paragraph in document.paragraphs:
            text = paragraph.text.strip()

            if not text:
                continue

            parts.append(text)

            if len("\n".join(parts)) > max_chars:
                break

        text_excerpt = "\n".join(parts)
        limited_text, truncated, original_length = _truncate_text(
            text_excerpt,
            max_chars,
        )

        result["text_excerpt"] = limited_text
        result["text_length_preview_source"] = original_length
        result["paragraph_count_previewed"] = len(parts)
        result["limits_applied"] = {
            "max_document_text_chars": max_chars,
            "text_truncated": truncated,
        }

        return result

    except Exception as exc:
        result["error"] = str(exc)
        return result


def _read_pdf_preview(
    file_path: str,
    *,
    max_pages: int = MAX_PDF_PREVIEW_PAGES,
    max_chars: int = MAX_DOCUMENT_TEXT_CHARS,
) -> dict[str, Any]:
    path = Path(file_path)

    result: dict[str, Any] = {
        "kind": "pdf_text_preview",
        "file_path": str(path),
        "exists": path.exists(),
        "text_excerpt": "",
        "pages_count": None,
        "pages_previewed": 0,
        "limits_applied": {
            "max_pdf_preview_pages": max_pages,
            "max_document_text_chars": max_chars,
            "text_truncated": False,
            "pages_truncated": False,
        },
        "error": None,
    }

    if not path.exists():
        result["error"] = "file_not_found"
        return result

    try:
        from pypdf import PdfReader
    except Exception as exc:
        result["error"] = f"pypdf_not_available: {exc}"
        return result

    try:
        reader = PdfReader(str(path))
        parts: list[str] = []

        page_count = len(reader.pages)
        pages_to_read = min(page_count, max_pages)
        pages_previewed = 0

        for page_index in range(pages_to_read):
            page = reader.pages[page_index]
            pages_previewed = page_index + 1

            try:
                text = page.extract_text() or ""
            except Exception:
                text = ""

            text = text.strip()

            if text:
                parts.append(f"--- page {page_index + 1} ---\n{text}")

            if len("\n\n".join(parts)) > max_chars:
                break

        text_excerpt = "\n\n".join(parts)
        limited_text, truncated, original_length = _truncate_text(
            text_excerpt,
            max_chars,
        )

        result["text_excerpt"] = limited_text
        result["text_length_preview_source"] = original_length
        result["pages_count"] = page_count
        result["pages_previewed"] = pages_previewed
        result["limits_applied"] = {
            "max_pdf_preview_pages": max_pages,
            "max_document_text_chars": max_chars,
            "text_truncated": truncated,
            "pages_truncated": page_count > pages_previewed,
        }

        if not result["text_excerpt"]:
            result["error"] = "no_text_layer_or_scanned_pdf"

        return result

    except Exception as exc:
        result["error"] = str(exc)
        return result


def _collect_preview_columns_and_text(
    preview: dict[str, Any],
) -> tuple[list[str], list[str]]:
    columns: list[str] = []
    text_parts: list[str] = []

    kind = preview.get("kind")

    if kind == "csv_table_preview":
        columns.extend(preview.get("columns") or [])

        for row in preview.get("sample_rows") or []:
            text_parts.append(json.dumps(row, ensure_ascii=False))

    elif kind in {"xlsx_table_preview", "xls_table_preview"}:
        for sheet in preview.get("sheets") or []:
            columns.extend(sheet.get("columns") or [])

            for row in sheet.get("sample_rows") or []:
                text_parts.append(json.dumps(row, ensure_ascii=False))

    elif kind in {"docx_text_preview", "pdf_text_preview"}:
        text_parts.append(preview.get("text_excerpt") or "")

    return columns, text_parts


def _detect_attachment_signals(preview: dict[str, Any]) -> dict[str, Any]:
    columns, text_parts = _collect_preview_columns_and_text(preview)

    normalized_columns: set[str] = set()

    for column in columns:
        text = str(column).strip().lower()

        if not text:
            continue

        if text.replace(".", "", 1).isdigit():
            continue

        if len(text) > 80:
            continue

        normalized_columns.add(text)

    full_text = "\n".join(text_parts).lower()

    def has_column(*names: str) -> bool:
        return any(name.lower() in normalized_columns for name in names)

    def has_text(*terms: str) -> bool:
        return any(term.lower() in full_text for term in terms)

    event_id_present = (
        has_column("event_id", "downtime_event_id", "id события")
        or has_text("event_id", "id события", "идентификатор события")
    )

    line_id_present = (
        has_column("line_id", "production_line_id", "id линии")
        or has_text("line_id", "id линии", "линия")
    )

    equipment_id_present = (
        has_column("equipment_id", "asset_id", "machine_id", "id оборудования")
        or has_text("equipment_id", "id оборудования", "оборудование")
    )

    start_timestamp_present = (
        has_column("downtime_start", "start_time", "started_at", "начало простоя")
        or has_text("downtime_start", "start_time", "начало простоя")
    )

    end_timestamp_present = (
        has_column("downtime_end", "end_time", "ended_at", "конец простоя")
        or has_text("downtime_end", "end_time", "конец простоя")
    )

    duration_present = (
        has_column("duration_min", "duration_minutes", "downtime_minutes", "длительность")
        or has_text("duration_min", "duration_minutes", "длительность простоя")
    )

    reason_present = (
        has_column("reason_code", "reason_description", "downtime_reason", "причина простоя")
        or has_text("reason_code", "reason_description", "причина простоя", "причины простоев")
    )

    cost_present = (
        has_column("cost_per_hour_rub", "estimated_loss_rub", "loss_rub", "стоимость часа")
        or has_text("cost_per_hour_rub", "estimated_loss_rub", "стоимость часа", "baseline", "потери")
    )

    downtime_log_detected = any(
        [
            event_id_present,
            start_timestamp_present,
            end_timestamp_present,
            duration_present,
            reason_present,
        ]
    )

    all_minimum_d002_fields_present = all(
        [
            event_id_present,
            line_id_present,
            equipment_id_present,
            start_timestamp_present,
            end_timestamp_present,
            duration_present,
            reason_present,
        ]
    )

    return {
        "downtime_log_detected": downtime_log_detected,
        "event_id_present": event_id_present,
        "line_id_present": line_id_present,
        "equipment_id_present": equipment_id_present,
        "start_timestamp_present": start_timestamp_present,
        "end_timestamp_present": end_timestamp_present,
        "duration_present": duration_present,
        "reason_present": reason_present,
        "cost_present": cost_present,
        "all_minimum_d002_fields_present": all_minimum_d002_fields_present,
        "detected_columns": sorted(normalized_columns),
    }


def _build_skipped_preview(
    *,
    file_path: str,
    exists: bool,
    reason: str,
    details: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "kind": "attachment_preview_skipped",
        "file_path": file_path,
        "exists": exists,
        "error": reason,
        "details": details or {},
        "attachment_review_policy": _build_attachment_limits_summary(),
    }


def _build_attachment_data_previews(
    attachments: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    previews: list[dict[str, Any]] = []
    preview_size_budget_used = 0

    for index, attachment in enumerate(attachments):
        original_filename = attachment.get("original_filename") or ""
        file_path = attachment.get("file_path") or ""
        resolved_path = _resolve_existing_path(file_path)
        suffix = Path(original_filename or file_path).suffix.lower()

        item = _make_attachment_item(
            attachment=attachment,
            resolved_path=resolved_path,
            suffix=suffix,
        )

        size_bytes = item.get("file_size_bytes")
        exists = bool(item.get("exists"))

        if index >= MAX_ATTACHMENTS_PER_BRIEF:
            preview = _build_skipped_preview(
                file_path=str(resolved_path),
                exists=exists,
                reason="attachments_count_limit_reached",
                details={
                    "max_attachments_per_brief": MAX_ATTACHMENTS_PER_BRIEF,
                    "attachment_index": index,
                },
            )

        elif not file_path:
            preview = _build_skipped_preview(
                file_path="",
                exists=False,
                reason="attachment_file_path_is_empty",
            )

        elif not exists:
            preview = _build_skipped_preview(
                file_path=str(resolved_path),
                exists=False,
                reason="file_not_found",
            )

        elif size_bytes is not None and size_bytes > MAX_ATTACHMENT_SIZE_BYTES:
            preview = _build_skipped_preview(
                file_path=str(resolved_path),
                exists=True,
                reason="single_attachment_size_limit_exceeded",
                details={
                    "file_size_bytes": size_bytes,
                    "file_size_mb": _bytes_to_mb(size_bytes),
                    "max_attachment_size_mb": MAX_ATTACHMENT_SIZE_MB,
                },
            )

        elif (
            size_bytes is not None
            and preview_size_budget_used + size_bytes > MAX_TOTAL_ATTACHMENTS_SIZE_BYTES
        ):
            preview = _build_skipped_preview(
                file_path=str(resolved_path),
                exists=True,
                reason="total_attachment_size_limit_exceeded",
                details={
                    "file_size_bytes": size_bytes,
                    "file_size_mb": _bytes_to_mb(size_bytes),
                    "current_total_preview_size_bytes": preview_size_budget_used,
                    "max_total_attachments_size_mb": MAX_TOTAL_ATTACHMENTS_SIZE_MB,
                },
            )

        elif suffix in {".xlsx", ".xlsm"}:
            preview_size_budget_used += size_bytes or 0
            preview = _read_xlsx_preview(str(resolved_path))

        elif suffix == ".xls":
            preview_size_budget_used += size_bytes or 0
            preview = _read_xls_preview(str(resolved_path))

        elif suffix == ".csv":
            preview_size_budget_used += size_bytes or 0
            preview = _read_csv_preview(str(resolved_path))

        elif suffix == ".docx":
            preview_size_budget_used += size_bytes or 0
            preview = _read_docx_preview(str(resolved_path))

        elif suffix == ".pdf":
            preview_size_budget_used += size_bytes or 0
            preview = _read_pdf_preview(str(resolved_path))

        else:
            preview = {
                "kind": "unsupported_preview",
                "file_path": str(resolved_path),
                "exists": exists,
                "error": "unsupported_file_type",
                "supported_formats": sorted(SUPPORTED_ATTACHMENT_PREVIEW_SUFFIXES),
            }

        item["preview"] = preview
        item["detected_signals"] = _detect_attachment_signals(preview)

        previews.append(item)

    return previews


def _compact_attachment_item_for_prompt(
    item: dict[str, Any],
    *,
    reason: str,
    max_excerpt_chars: int,
) -> dict[str, Any]:
    preview = item.get("preview") or {}
    serialized = json.dumps(
        item,
        ensure_ascii=False,
        default=str,
    )

    excerpt = _limit_text(serialized, max(max_excerpt_chars, 0))

    return {
        "attachment_id": item.get("attachment_id"),
        "input_pack_id": item.get("input_pack_id"),
        "original_filename": item.get("original_filename"),
        "stored_filename": item.get("stored_filename"),
        "file_type": item.get("file_type"),
        "uploaded_at": item.get("uploaded_at"),
        "suffix": item.get("suffix"),
        "exists": item.get("exists"),
        "file_size_bytes": item.get("file_size_bytes"),
        "file_size_mb": item.get("file_size_mb"),
        "detected_signals": item.get("detected_signals"),
        "preview_summary": {
            "kind": preview.get("kind"),
            "exists": preview.get("exists"),
            "error": preview.get("error"),
            "limits_applied": preview.get("limits_applied"),
        },
        "preview_truncated_for_prompt": True,
        "prompt_truncation_reason": reason,
        "serialized_preview_excerpt": excerpt,
    }


def _limit_attachment_prompt_blocks(
    previews: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    limited_previews: list[dict[str, Any]] = []
    used_chars = 0

    for item in previews:
        remaining_total_chars = MAX_TOTAL_ATTACHMENT_PROMPT_CHARS - used_chars

        if remaining_total_chars <= 0:
            limited_previews.append(
                {
                    "kind": "attachment_preview_skipped",
                    "reason": "total_attachment_prompt_limit_reached",
                    "limit_chars": MAX_TOTAL_ATTACHMENT_PROMPT_CHARS,
                    "attachment_review_policy": _build_attachment_limits_summary(),
                }
            )
            break

        serialized = json.dumps(
            item,
            ensure_ascii=False,
            default=str,
        )

        per_item_limit = min(
            MAX_SINGLE_ATTACHMENT_PROMPT_CHARS,
            remaining_total_chars,
        )

        if len(serialized) > per_item_limit:
            item_for_prompt = _compact_attachment_item_for_prompt(
                item,
                reason="single_or_total_attachment_prompt_limit_applied",
                max_excerpt_chars=per_item_limit,
            )
        else:
            item_for_prompt = item

        item_serialized = json.dumps(
            item_for_prompt,
            ensure_ascii=False,
            default=str,
        )

        if len(item_serialized) > remaining_total_chars:
            item_for_prompt = _compact_attachment_item_for_prompt(
                item,
                reason="total_attachment_prompt_limit_applied",
                max_excerpt_chars=remaining_total_chars,
            )
            item_serialized = json.dumps(
                item_for_prompt,
                ensure_ascii=False,
                default=str,
            )

        limited_previews.append(item_for_prompt)
        used_chars += min(len(item_serialized), remaining_total_chars)

    return limited_previews


def _build_attachment_evidence_summary(
    attachment_data_previews: list[dict[str, Any]],
    attachment_inventory: dict[str, Any] | None = None,
) -> dict[str, Any]:
    field_sources: dict[str, list[str]] = {
        "event_id": [],
        "line_id": [],
        "equipment_id": [],
        "start_timestamp": [],
        "end_timestamp": [],
        "duration": [],
        "downtime_reason": [],
        "cost_or_loss": [],
    }

    downtime_log_files: list[str] = []
    technical_regulation_files: list[str] = []
    commission_act_files: list[str] = []
    detected_columns_by_file: list[dict[str, Any]] = []
    files_reviewed: list[str] = []
    files_skipped: list[dict[str, Any]] = []

    for item in attachment_data_previews:
        filename = (
            item.get("original_filename")
            or item.get("stored_filename")
            or f"attachment_{item.get('attachment_id')}"
        )

        preview = item.get("preview") or {}
        signals = item.get("detected_signals") or {}
        preview_kind = preview.get("kind")
        preview_error = preview.get("error")

        if preview_kind == "attachment_preview_skipped":
            files_skipped.append(
                {
                    "filename": filename,
                    "reason": preview_error,
                    "details": preview.get("details") or {},
                }
            )
        elif preview_error:
            files_skipped.append(
                {
                    "filename": filename,
                    "reason": preview_error,
                    "details": {},
                }
            )
        else:
            files_reviewed.append(filename)

        if signals.get("downtime_log_detected"):
            downtime_log_files.append(filename)

        if signals.get("event_id_present"):
            field_sources["event_id"].append(filename)

        if signals.get("line_id_present"):
            field_sources["line_id"].append(filename)

        if signals.get("equipment_id_present"):
            field_sources["equipment_id"].append(filename)

        if signals.get("start_timestamp_present"):
            field_sources["start_timestamp"].append(filename)

        if signals.get("end_timestamp_present"):
            field_sources["end_timestamp"].append(filename)

        if signals.get("duration_present"):
            field_sources["duration"].append(filename)

        if signals.get("reason_present"):
            field_sources["downtime_reason"].append(filename)

        if signals.get("cost_present"):
            field_sources["cost_or_loss"].append(filename)

        detected_columns = signals.get("detected_columns") or []

        if detected_columns:
            detected_columns_by_file.append(
                {
                    "filename": filename,
                    "detected_columns": detected_columns,
                }
            )

        text_excerpt = ""

        if preview.get("kind") in {"docx_text_preview", "pdf_text_preview"}:
            text_excerpt = (preview.get("text_excerpt") or "").lower()

        if any(
            term in text_excerpt
            for term in [
                "технический регламент",
                "регламент фиксации",
                "регламент остановок",
                "порядок фиксации",
            ]
        ):
            technical_regulation_files.append(filename)

        if any(
            term in text_excerpt
            for term in [
                "акт комиссии",
                "комиссия",
                "расследование причин",
                "выводы комиссии",
            ]
        ):
            commission_act_files.append(filename)

    minimum_downtime_export_fields_present = all(
        [
            bool(field_sources["event_id"]),
            bool(field_sources["line_id"]),
            bool(field_sources["equipment_id"]),
            bool(field_sources["start_timestamp"]),
            bool(field_sources["end_timestamp"]),
            bool(field_sources["duration"]),
            bool(field_sources["downtime_reason"]),
        ]
    )

    return {
        "files_count": len(attachment_data_previews),
        "files_reviewed": sorted(set(files_reviewed)),
        "files_skipped": files_skipped,
        "attachment_inventory": attachment_inventory or {},
        "attachment_review_policy": _build_attachment_limits_summary(),
        "downtime_log_files": sorted(set(downtime_log_files)),
        "technical_regulation_files": sorted(set(technical_regulation_files)),
        "commission_act_files": sorted(set(commission_act_files)),
        "field_sources": {
            key: sorted(set(value))
            for key, value in field_sources.items()
        },
        "minimum_downtime_export_fields_present": minimum_downtime_export_fields_present,
        "detected_columns_by_file": detected_columns_by_file,
        "d001_interpretation_rules": {
            "attachment_review_mode": [
                "Uploaded files are used only in limited consulting-grade review mode.",
                "Use attachment previews for data structure, field presence, MVP readiness and baseline requirements.",
                "Do not claim full BI analytics, deep sensor analytics, production model training or final ROI calculation.",
            ],
            "if_minimum_downtime_export_fields_present": [
                "Do not say that event_id, line_id, equipment_id, start/end timestamps, duration or downtime reason are missing.",
                "Treat identifiers and timestamps as present in the test export.",
                "Remaining questions should be about representativeness, stability of the export format, data quality, timezone, completeness, security approval and baseline confirmation.",
                "Data readiness may remain MEDIUM because the export is test evidence, but missing-data wording must not contradict the attachment evidence.",
            ],
            "economics_rule": [
                "If cost_per_hour_rub or estimated_loss_rub is present, treat economy as partially evidenced.",
                "Still require finance confirmation of cost per downtime hour, monthly baseline and target effect.",
            ],
            "pdf_word_rule": [
                "Use PDF/DOCX text as evidence for commission conclusions, technical regulation, downtime cause classification and process documentation.",
                "If technical_regulation_files is not empty, treat the attachments as evidence that a technical regulation / технический регламент or equivalent downtime recording procedure exists.",
                "Do not treat PDF/DOCX as structured event logs unless table fields are explicitly detected.",
            ],
        },
    }


def _replace_markdown_table_row(
    markdown: str,
    row_label: str,
    replacement_row: str,
) -> str:
    lines = markdown.splitlines()

    for index, line in enumerate(lines):
        if line.lstrip().startswith("|") and row_label in line:
            lines[index] = replacement_row

    return "\n".join(lines)


def _collect_evidence_files(evidence_summary: dict[str, Any]) -> str:
    files: set[str] = set()

    for filename in evidence_summary.get("downtime_log_files", []) or []:
        files.add(str(filename))

    for filename in evidence_summary.get("commission_act_files", []) or []:
        files.add(str(filename))

    for filenames in (evidence_summary.get("field_sources", {}) or {}).values():
        for filename in filenames or []:
            files.add(str(filename))

    if not files:
        return "приложенные файлы"

    return ", ".join(sorted(files))


def _postprocess_d001_with_attachment_evidence(
    markdown: str,
    evidence_summary: dict[str, Any] | None,
) -> str:
    """
    Детерминированная защита D-001 от противоречия с evidence summary.

    Если вложения содержат минимальный набор полей downtime export,
    D-001 не должен писать, что event_id, line_id, equipment_id,
    start/end timestamps, duration или downtime reason отсутствуют.
    """

    if not evidence_summary:
        return markdown

    if not evidence_summary.get("minimum_downtime_export_fields_present"):
        return markdown

    evidence_files = _collect_evidence_files(evidence_summary)

    result = markdown

    result = _replace_markdown_table_row(
        result,
        "Главный вывод",
        "| Главный вывод           | Проект можно продолжать с ограничениями: минимальный набор идентификаторов и временных меток представлен в тестовой выгрузке; требуется подтвердить репрезентативность выгрузки, стабильность ID, timezone/формат timestamp, качество данных, baseline и правила ИБ. |",
    )

    result = _replace_markdown_table_row(
        result,
        "Идентификаторы объектов",
        f"| Идентификаторы объектов     | YES_WITH_VALIDATION           | event_id, line_id, equipment_id представлены в тестовой выгрузке и приложенных материалах: {evidence_files} | Подтвердить стабильность ID в реальных системах; ID наряда ремонта — опционально, если требуется связка с ТОиР | LOW/MEDIUM       |",
    )

    result = _replace_markdown_table_row(
        result,
        "Временные метки",
        f"| Временные метки             | YES_WITH_VALIDATION           | start_time, end_time, duration_min представлены в тестовой выгрузке и приложенных материалах: {evidence_files} | Подтвердить timezone, единый формат timestamp, полноту периода и отсутствие систематических пропусков | LOW/MEDIUM       |",
    )

    result = _replace_markdown_table_row(
        result,
        "MUST      | Подтвердить ID события простоя",
        "| MUST      | Подтвердить стабильность ID события простоя в реальных системах | Для проверки связности выгрузки при промышленном использовании | Руководитель производства / владелец данных |",
    )

    result = _replace_markdown_table_row(
        result,
        "MUST      | Уточнить временные метки событий",
        "| MUST      | Подтвердить timezone, формат timestamp и полноту периода выгрузки | Для корректного расчёта длительности простоев и сопоставления событий | Руководитель производства / владелец данных |",
    )

    replacements = {
        "Проект можно продолжать с ограничениями, требуется уточнение идентификаторов и временных меток, а также согласование правил ИБ.": (
            "Проект можно продолжать с ограничениями: минимальный набор идентификаторов и временных меток представлен в тестовой выгрузке; требуется подтвердить репрезентативность выгрузки, стабильность ID, timezone/формат timestamp, качество данных, baseline и правила ИБ."
        ),
        "требуется уточнение идентификаторов и временных меток": (
            "требуется подтверждение репрезентативности выгрузки, стабильности ID, timezone/формата timestamp и правил ИБ"
        ),
        "необходимо уточнить идентификаторы объектов и временные метки": (
            "необходимо подтвердить репрезентативность выгрузки, стабильность ID, timezone/формат timestamp и качество данных"
        ),
        "необходимо уточнить идентификаторы объектов и временные метки, а также согласовать правила ИБ": (
            "необходимо подтвердить репрезентативность выгрузки, стабильность ID, timezone/формат timestamp, качество данных и согласовать правила ИБ"
        ),
        "получить ограниченную обезличенную выгрузку данных, проверить идентификаторы и временные метки": (
            "использовать ограниченную обезличенную тестовую выгрузку данных, подтвердить стабильность идентификаторов, timezone/формат timestamp и полноту периода"
        ),
        "ID события, ID наряда ремонта": (
            "стабильность ID события в реальных системах; ID наряда ремонта — опционально для связки с ТОиР"
        ),
        "Начало и конец простоя | Временные метки событий": (
            "start_time, end_time, duration_min | Подтвердить timezone, формат timestamp и полноту периода"
        ),
        "Подтвердить ID события простоя": (
            "Подтвердить стабильность ID события простоя в реальных системах"
        ),
        "Уточнить временные метки событий": (
            "Подтвердить timezone, формат timestamp и полноту периода выгрузки"
        ),
        "Доступны данные о событиях простоев, частично подтверждены ID и временные метки": (
            "Тестовая выгрузка и приложенные материалы содержат события простоев, event_id, line_id, equipment_id, start_time, end_time, duration_min и причины простоев; требуется подтвердить репрезентативность, стабильность формата и качество данных"
        ),
        "частично подтверждены ID и временные метки": (
            "представлены ID и временные метки в тестовой выгрузке; требуется подтверждение стабильности формата"
        ),
        "необходимость подтверждения ID, временных меток и согласования правил ИБ": (
            "необходимость подтверждения стабильности ID, timezone/формата timestamp, полноты периода, качества данных и согласования правил ИБ"
        ),
        "подтверждение наличия всех обязательных идентификаторов и временных меток": (
            "подтверждение стабильности обязательных идентификаторов, timezone/формата timestamp, полноты периода и качества данных"
        ),
        "ID оборудования, ID линии, ID события и временные метки относятся к готовности данных": (
            "ID оборудования, ID линии, ID события и временные метки представлены в тестовой выгрузке; для D-002 требуется подтвердить стабильность формата, полноту периода и качество данных"
        ),
        "Однако необходимо уточнить идентификаторы объектов и временные метки, а также согласовать правила ИБ.": (
            "Минимальные идентификаторы и временные метки представлены в тестовой выгрузке; необходимо подтвердить репрезентативность, стабильность ID, timezone/формат timestamp, качество данных и согласовать правила ИБ."
        ),
        "Необходимо подтвердить ID события, оборудования и линии, а также временные метки.": (
            "ID события, оборудования, линии и временные метки представлены в тестовой выгрузке; необходимо подтвердить стабильность ID, timezone/формат timestamp и полноту периода."
        ),
        "Невозможность связать событие простоя с оборудованием, линией, временной меткой, длительностью и причиной в единой выгрузке.": (
            "Риск нестабильной связки события простоя с оборудованием, линией, временной меткой, длительностью и причиной при переходе от тестовой выгрузки к промышленному регулярному экспорту."
        ),
    }

    for old_text, new_text in replacements.items():
        result = result.replace(old_text, new_text)

    # Не завышаем уверенность в экономике и KPI.
    result = result.replace(
        "| Подтверждённые числа  | 80–120 событий в месяц | HIGH |  |",
        "| Предварительно указанные числа | 80–120 событий в месяц | MEDIUM | Требует подтверждения на baseline-периоде |",
    )

    result = result.replace(
        "| Подтвержденные числа  | 80–120 событий в месяц | HIGH |  |",
        "| Предварительно указанные числа | 80–120 событий в месяц | MEDIUM | Требует подтверждения на baseline-периоде |",
    )

    result = result.replace(
        "| KPI можно измерить          | YES                          | Снижение простоев на 5–10% | —               | —                |",
        "| KPI можно измерить          | YES_WITH_VALIDATION          | Гипотезу KPI можно проверить после фиксации baseline, периода сравнения и правил расчёта | Подтвердить baseline и критерии успеха | MEDIUM           |",
    )

    result = result.replace(
        "| Качество данных понятно     | PARTIAL                      | Дубликаты, пропуски | Требует уточнения | Влияет на анализ |",
        "| Качество данных понятно     | PARTIAL                      | Структура видна на тестовой выгрузке | Проверить дубликаты, пропуски, стабильность формата и полноту периода | Влияет на D-002 и MVP |",
    )

    result = result.replace(
        "| Потенциальный эффект  | 5–10% снижение простоев | MEDIUM |  |",
        "| Потенциальный эффект  | Предварительная гипотеза снижения управляемых потерь | MEDIUM | Конкретный эффект фиксируется после baseline |",
    )

    result = result.replace(
        "| KPI успеха | Проверить гипотезу снижения простоев на 5–10% и сокращения ручной отчетности на 20–30%. |",
        "| KPI успеха | Проверить гипотезу снижения управляемых потерь и сокращения ручной работы; конкретные целевые значения фиксируются после baseline. |",
    )

    result = result.replace(
        "| KPI успеха | Проверить гипотезу снижения простоев на 5–10% и сокращения ручной отчётности на 20–30%. |",
        "| KPI успеха | Проверить гипотезу снижения управляемых потерь и сокращения ручной работы; конкретные целевые значения фиксируются после baseline. |",
    )

    result = result.replace(
        "Интеграции с MES и WMS отсутствуют, что также требует внимания.",
        "Интеграции с MES/SCADA/API не обязательны для первого MVP и могут быть отложены до подтверждения ценности пилота.",
    )

    result = result.replace(
        "Интеграции с MES и WMS отсутствуют",
        "Интеграции с MES/SCADA/API не обязательны для первого MVP",
    )

    evidence_note = f"""
> Комментарий по приложенным материалам: загруженные файлы использованы в режиме limited consulting-grade review. Они содержат минимальный набор полей для анализа простоев: event_id, line_id, equipment_id, start_time, end_time, duration_min и reason_code/reason_description. Источники: {evidence_files}. Поэтому D-001 считает эти поля представленными в тестовой выгрузке; оставшиеся ограничения относятся к репрезентативности, стабильности формата, качеству данных, timezone, baseline и правилам ИБ. Глубокая BI-аналитика, production-моделирование и финальный расчёт ROI по файлам не выполнялись.
""".strip()

    if "Комментарий по приложенным материалам:" not in result and "## 2. Что уже известно" in result:
        result = result.replace(
            "## 2. Что уже известно",
            f"{evidence_note}\n\n---\n\n## 2. Что уже известно",
            1,
        )

    return result


def get_latest_input_pack_for_d001(diagnostic_run_id: int) -> dict[str, Any]:
    input_pack_row = _get_latest_input_pack_row(diagnostic_run_id)
    input_pack = dict(input_pack_row)

    raw_payload = _load_json_object(input_pack.get("raw_payload"), "raw_payload")

    _validate_raw_payload_for_d001(
        input_pack_id=int(input_pack["id"]),
        raw_payload=raw_payload,
    )

    attachments = _get_attachments_for_input_pack(int(input_pack["id"]))
    attachment_inventory = _build_attachment_inventory(attachments)
    attachment_data_previews = _build_attachment_data_previews(attachments)
    attachment_evidence_summary = _build_attachment_evidence_summary(
        attachment_data_previews,
        attachment_inventory,
    )

    return {
        "input_pack": input_pack,
        "raw_payload": raw_payload,
        "attachments": attachments,
        "attachment_inventory": attachment_inventory,
        "attachment_data_previews": attachment_data_previews,
        "attachment_data_previews_for_prompt": _limit_attachment_prompt_blocks(
            attachment_data_previews
        ),
        "attachment_evidence_summary": attachment_evidence_summary,
    }


def ensure_normalized_payload(
    diagnostic_run_id: int,
    *,
    input_pack: dict[str, Any] | None = None,
    raw_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if input_pack is None or raw_payload is None:
        data = get_latest_input_pack_for_d001(diagnostic_run_id)
        input_pack = data["input_pack"]
        raw_payload = data["raw_payload"]

    if raw_payload.get("brief_type") == "industrial_ai":
        return raw_payload

    if input_pack.get("normalized_payload"):
        return _load_json_object(
            input_pack["normalized_payload"],
            "normalized_payload",
        )

    return normalize_input_pack(int(input_pack["id"]))


def get_existing_d001_result(diagnostic_run_id: int) -> str | None:
    diagnostic_run = _get_diagnostic_run_row(diagnostic_run_id)
    return diagnostic_run["d001_result"] or None


def build_d001_prompt_input(
    diagnostic_run_id: int,
    *,
    data: dict[str, Any] | None = None,
) -> str:
    diagnostic_run = _get_diagnostic_run_row(diagnostic_run_id)

    if data is None:
        data = get_latest_input_pack_for_d001(diagnostic_run_id)

    input_pack = data["input_pack"]
    raw_payload = data["raw_payload"]
    normalized_payload = ensure_normalized_payload(
        diagnostic_run_id,
        input_pack=input_pack,
        raw_payload=raw_payload,
    )

    safe_diagnostic_run = _sanitize_diagnostic_run(diagnostic_run)

    safe_input_pack = {
        "id": input_pack.get("id"),
        "diagnostic_run_id": input_pack.get("diagnostic_run_id"),
        "status": input_pack.get("status"),
        "created_at": input_pack.get("created_at"),
        "updated_at": input_pack.get("updated_at"),
        "brief_type": raw_payload.get("brief_type"),
        "brief_version": raw_payload.get("brief_version"),
        "source": raw_payload.get("source"),
        "submitted_at": raw_payload.get("submitted_at"),
        "raw_payload_non_empty": _count_non_empty(raw_payload),
        "industrial_ai_non_empty": _count_non_empty(
            raw_payload.get("industrial_ai", {})
        ),
    }

    attachment_review_policy = _build_attachment_limits_summary()

    return f"""
# INPUT FOR D-001 DIAGNOSTIC ASSESSMENT AGENT

## Diagnostic Run

{json.dumps(safe_diagnostic_run, ensure_ascii=False, indent=2)}

## Selected Input Pack

{json.dumps(safe_input_pack, ensure_ascii=False, indent=2)}

## Raw Diagnostic Input Pack / Industrial AI Brief

This is the source of truth.

If `brief_type` is `industrial_ai`, analyze the `industrial_ai` object directly.
Do not treat the brief as empty if `industrial_ai` contains filled fields.
Do not let the normalized helper payload override the raw Industrial AI Brief.

{json.dumps(raw_payload, ensure_ascii=False, indent=2)}

## Normalized Diagnostic Input Pack

Use this only as a helper.
If this section conflicts with the raw payload above, the raw payload wins.

{json.dumps(normalized_payload, ensure_ascii=False, indent=2)}

## Attachment Review Policy

Uploaded files are used in limited consulting-grade review mode.

Use attachments for:
- checking data structure;
- detecting key fields;
- evaluating MVP readiness;
- forming baseline, identifier, regular export and data-quality requirements.

Do not use attachments for:
- full BI analytics;
- deep sensor analytics;
- production model training;
- final ROI calculation;
- legal or audit-grade conclusions.

{json.dumps(attachment_review_policy, ensure_ascii=False, indent=2)}

## Attachments Inventory

{json.dumps(data.get("attachment_inventory", {}), ensure_ascii=False, indent=2)}

## Attachments Summary

{json.dumps(data.get("attachments", []), ensure_ascii=False, indent=2)}

## Attachment Evidence Summary

This section is a deterministic summary of parsed attachment evidence.

Use this section as the primary evidence layer for deciding whether critical data fields
are present in uploaded files.

If `minimum_downtime_export_fields_present` is true:
- do not say that event_id is missing;
- do not say that line_id is missing;
- do not say that equipment_id is missing;
- do not say that start/end timestamps are missing;
- do not say that duration is missing;
- do not say that downtime reason is missing.

You may still require confirmation of:
- export representativeness;
- stability of IDs across real systems;
- timezone and timestamp format;
- duplicates and missing values;
- completeness of the period;
- security approval;
- finance-confirmed baseline.

{json.dumps(data.get("attachment_evidence_summary", {}), ensure_ascii=False, indent=2)}

## Attachments Data Preview

This section contains limited parsed previews of supported attachments:
Excel/XLSX/XLSM/XLS, CSV, Word/DOCX and PDF.

For Excel/CSV files, use columns and sample rows as evidence of data availability.
For Word/PDF files, use extracted text as evidence of documented process, constraints,
baseline, security requirements, data field descriptions or export specification.

The full files are stored, but this prompt receives only compact previews and evidence summaries.
Respect the limits in Attachment Review Policy.

If the preview contains event_id, line_id, equipment_id, start/end timestamps,
duration and downtime reason fields, do not state that these fields are missing.
You may still require confirmation that the export is representative, complete,
stable, approved by the client and safe to process.

{json.dumps(data.get("attachment_data_previews_for_prompt", []), ensure_ascii=False, indent=2)}
""".strip()


def run_d001_diagnostic_assessment(
    diagnostic_run_id: int,
    force_rebuild: bool = False,
) -> str:
    existing_result = get_existing_d001_result(diagnostic_run_id)

    if existing_result and not force_rebuild:
        return existing_result

    d001_data = get_latest_input_pack_for_d001(diagnostic_run_id)

    update_diagnostic_status(
        diagnostic_run_id=diagnostic_run_id,
        status=DIAGNOSTIC_STATUS_D001_RUNNING,
    )

    prompt_input = build_d001_prompt_input(
        diagnostic_run_id,
        data=d001_data,
    )

    result = run_agent_with_prompt(
        agent_prompt_name=D001_AGENT_PROMPT_NAME,
        user_input=prompt_input,
    )

    result = _postprocess_d001_with_attachment_evidence(
        markdown=result,
        evidence_summary=d001_data.get("attachment_evidence_summary", {}),
    )

    save_d001_result(
        diagnostic_run_id=diagnostic_run_id,
        result=result,
    )

    update_diagnostic_status(
        diagnostic_run_id=diagnostic_run_id,
        status=DIAGNOSTIC_STATUS_D001_COMPLETED,
    )

    return result